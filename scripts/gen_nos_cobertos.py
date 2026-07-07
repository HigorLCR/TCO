"""
Cria um xlsx do zero com a pagina 'Nos_AST_Cobertos', colorida conforme a
cobertura de nos AST das funcoes de recursive_functions/benchmark.

Linhas  = funcoes (as do node_matrix.txt), na ordem historica + faltantes.
Colunas = nos AST (cabecalho 'ast.<No>', ordem ASDL fixa abaixo).

Regra de cor (por coluna/no), de cima para baixo:
  VERDE  = primeira funcao que cobre o no (quem "identificou" o no).
  CINZA  = qualquer funcao ABAIXO da verde (no ja coberto por outra).
  BRANCO = funcoes acima da verde, ou no que ninguem cobre.

Gera um arquivo NOVO (nao depende de ling_rec.xlsx). Sem Tabela/ListObject
(o zebrado dela mascarava os preenchimentos manuais), com altura de linha
uniforme e coluna A sem realces antigos.

Uso:
    python scripts/gen_nos_cobertos.py
"""

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).parent.parent
SAIDA_DIR = BASE / "arquivos" / "xlsx"  # diretorio padrao de saida dos xlsx gerados
SAIDA = SAIDA_DIR / "nos_ast_cobertos.xlsx"
MATRIX = BASE / "arquivos" / "txt" / "node_matrix.txt"

ABA = "Nós_AST_Cobertos"
ALTURA = 22.5

VERDE = PatternFill(start_color="FF70AD47", end_color="FF70AD47", fill_type="solid")
CINZA = PatternFill(start_color="FFA6A6A6", end_color="FFA6A6A6", fill_type="solid")
BRANCO = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

HDR_FILL = PatternFill(start_color="FF217346", end_color="FF217346", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFFFF")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
COLA_ALIGN = Alignment(horizontal="left", vertical="center")

# Ordem fixa dos nos AST (cabecalho 'ast.<No>'), em ordem ASDL, como na planilha.
NODE_HEADERS = [
    "ast.Module", "ast.FunctionDef", "ast.AsyncFunctionDef", "ast.ClassDef",
    "ast.Return", "ast.Delete", "ast.Assign", "ast.AugAssign", "ast.AnnAssign",
    "ast.TypeAlias", "ast.For", "ast.AsyncFor", "ast.While", "ast.If", "ast.With",
    "ast.AsyncWith", "ast.Match", "ast.Raise", "ast.Try", "ast.TryStar",
    "ast.Assert", "ast.Import", "ast.ImportFrom", "ast.Global", "ast.Nonlocal",
    "ast.Expr", "ast.Pass", "ast.Break", "ast.Continue", "ast.BoolOp",
    "ast.NamedExpr", "ast.BinOp", "ast.UnaryOp", "ast.Lambda", "ast.IfExp",
    "ast.Dict", "ast.Set", "ast.ListComp", "ast.SetComp", "ast.DictComp",
    "ast.GeneratorExp", "ast.Await", "ast.Yield", "ast.YieldFrom", "ast.Compare",
    "ast.Call", "ast.FormattedValue", "ast.JoinedStr", "ast.Constant",
    "ast.Attribute", "ast.Subscript", "ast.Starred", "ast.Name", "ast.List",
    "ast.Tuple", "ast.Slice", "ast.Load", "ast.Store", "ast.Del", "ast.Add",
    "ast.Sub", "ast.Mult", "ast.Div", "ast.Mod", "ast.Pow", "ast.FloorDiv",
    "ast.LShift", "ast.RShift", "ast.BitOr", "ast.BitXor", "ast.BitAnd",
    "ast.MatMult", "ast.Invert", "ast.Not", "ast.UAdd", "ast.USub", "ast.And",
    "ast.Or", "ast.Eq", "ast.NotEq", "ast.Lt", "ast.LtE", "ast.Gt", "ast.GtE",
    "ast.Is", "ast.IsNot", "ast.In", "ast.NotIn", "ast.MatchValue",
    "ast.MatchSingleton", "ast.MatchSequence", "ast.MatchMapping",
    "ast.MatchClass", "ast.MatchStar", "ast.MatchAs", "ast.MatchOr",
    "ast.TypeVar", "ast.ParamSpec", "ast.TypeVarTuple", "ast.arg",
    "ast.arguments", "ast.keyword", "ast.alias", "ast.withitem",
    "ast.comprehension", "ast.ExceptHandler", "ast.match_case",
]

# Ordem historica das funcoes na coluna A (nomes de exibicao originais).
DISPLAY_ORDER = [
    "Factorial()", "Fibonacci()", "Sum()", "mdc()", "woodcutter()",
    "Binary_Search_Tree()", "quick_select()", "linear_search()",
    "non_negative_int_contain_digit()", "Hannoi ()", "MergeSort()",
    "NumSearchTree ()", "fast_pow()", "count_bits()", "flatten()",
    "first_missing()", "is_valid_bst()", "fib_memo()", "countdown()",
    "safe_parse()", "sum_nested()", "remove_keys()", "find_first()",
    "count_matches()",
]

# nome de exibicao (coluna A) -> arquivo .py (stem) no node_matrix
DISPLAY_TO_STEM = {
    "Factorial()": "factorial",
    "Fibonacci()": "fibonacci",
    "Sum()": "sum",
    "mdc()": "mdc",
    "woodcutter()": "woodcutter",
    "Binary_Search_Tree()": "binary_search",
    "quick_select()": "quickselect",
    "linear_search()": "linear_search",
    "non_negative_int_contain_digit()": "non_negative_int_contain_digit",
    "Hannoi ()": "Hannoi",
    "MergeSort()": "MergeSort",
    "NumSearchTree ()": "NumSearchTree",
    "fast_pow()": "fast_pow",
    "count_bits()": "count_bits",
    "flatten()": "flatten",
    "first_missing()": "first_missing",
    "is_valid_bst()": "valid_bst",
    "fib_memo()": "fib_memo",
    "countdown()": "countdown",
    "safe_parse()": "safe_parse",
    "sum_nested()": "sum_nested",
    "remove_keys()": "remove_keys",
    "find_first()": "find_first",
    "count_matches()": "count_matches",
}


def carregar_cobertura() -> dict[str, set[str]]:
    """node_matrix.txt -> {stem_do_arquivo: set(nos cobertos)}."""
    text = MATRIX.read_bytes().decode("utf-16")
    linhas = [l for l in text.split("\r\n") if l]
    arquivos = linhas[0].split("|")[1:]  # ex.: 'factorial.py'
    cov = {a: set() for a in arquivos}
    for l in linhas[1:]:
        cels = l.split("|")
        no = cels[0]
        for i, v in enumerate(cels[1:]):
            if v == "X":
                cov[arquivos[i]].add(no)
    return {a[:-3]: s for a, s in cov.items()}  # tira '.py'


def stem_de(display: str) -> str:
    """Resolve o stem do arquivo a partir do nome de exibicao."""
    if display in DISPLAY_TO_STEM:
        return DISPLAY_TO_STEM[display]
    return display.replace("()", "").strip()  # nomes derivados (stem + '()')


def ordem_funcoes(cov: dict[str, set[str]]) -> list[tuple[str, str]]:
    """Lista (display, stem): ordem historica + faltantes (novas) ordenadas."""
    linhas = [(d, stem_de(d)) for d in DISPLAY_ORDER]
    ja = {s for _, s in linhas}
    for stem in sorted(s for s in cov if s not in ja):
        linhas.append((f"{stem}()", stem))
    return linhas


def main() -> None:
    cov = carregar_cobertura()
    funcoes = ordem_funcoes(cov)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ABA

    # cabecalho
    ws.cell(1, 1, "Função")
    col_de_no = {}
    for i, header in enumerate(NODE_HEADERS):
        c = i + 2
        ws.cell(1, c, header)
        col_de_no[header.replace("ast.", "")] = c

    # coluna A (funcoes) + altura uniforme
    linha_stem: dict[int, str] = {}
    for i, (display, stem) in enumerate(funcoes):
        r = i + 2
        cell = ws.cell(r, 1, display)
        cell.alignment = COLA_ALIGN
        linha_stem[r] = stem
    ultima = len(funcoes) + 1

    # coloracao por coluna/no
    sem_cobertura = []
    for no, c in col_de_no.items():
        achou = False
        for r in range(2, ultima + 1):
            tem = no in cov.get(linha_stem[r], set())
            cell = ws.cell(r, c)
            if not achou and tem:
                cell.fill = VERDE
                achou = True
            elif achou:
                cell.fill = CINZA
            else:
                cell.fill = BRANCO
        if not achou:
            sem_cobertura.append(no)

    # estilo do cabecalho
    for c in range(1, len(NODE_HEADERS) + 2):
        cell = ws.cell(1, c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = HDR_ALIGN

    # altura uniforme (cabecalho + dados) e largura da coluna A
    for r in range(1, ultima + 1):
        ws.row_dimensions[r].height = ALTURA
    ws.column_dimensions["A"].width = 30
    ws.freeze_panes = "B2"

    SAIDA_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(SAIDA)

    print(f"Arquivo criado : {SAIDA}")
    print(f"Aba            : {ABA}")
    print(f"Funcoes        : {len(funcoes)} (linhas 2..{ultima})")
    print(f"Nos (colunas)  : {len(NODE_HEADERS)}")
    novas = [d for d, _ in funcoes[len(DISPLAY_ORDER):]]
    print(f"Adicionadas    : {len(novas)} -> {', '.join(novas) if novas else '-'}")
    print(f"Sem cobertura ({len(sem_cobertura)}): {', '.join(sem_cobertura) if sem_cobertura else '-'}")


if __name__ == "__main__":
    main()
