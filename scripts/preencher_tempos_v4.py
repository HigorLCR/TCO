"""
Preenche a aba 'Tempos_Execucao_v4' de ling_rec.xlsx com os tempos coletados
em benchmark_results.csv e com os parametros reais usados nos benchmarks.

Layout da aba (colunas = funcoes, linhas = metricas):
    linha 2: Tempo_Recursao_Media        <- <nome>.py           (recursivo)
    linha 3: Tempo_Iteracao_Tail_Media   <- output_<nome>.py    (tail, so quem tem)
    linha 4: Tempo_Iteracao_Media        <- <nome>_nonrec.py    (nonrec)
    linha 5: Sobrecarga_Tail             = formula rec/tail   (por rotulo)
    linha 6: Sobrecarga                  = formula rec/nonrec (por rotulo)
    linha 7: Numero_Iteracoes            <- qtd_execucoes do benchmark
    linha 8: Parametros                  <- args reais no timeit(lambda: f(...))

Metrica de tempo: tempo_ms_por_chamada (media por chamada).
Numero_Iteracoes: qtd_execucoes (repeticoes do timeit) por funcao.
Parametros: argumentos da chamada no timeit, com variaveis de modulo resolvidas.
Valores ausentes (timeout / sem_timing) viram '-'.

As formulas de sobrecarga localizam as linhas de tempo pelo ROTULO (INDEX/MATCH
na coluna A), entao sao imunes a insercao/reordenacao de linhas.

Uso:
    python scripts/preencher_tempos_v4.py
"""

import ast
import csv
import re
import sys
from pathlib import Path

import openpyxl

BASE = Path(__file__).parent.parent
XLSX = BASE / "ling_rec.xlsx"
CSV = BASE / "benchmark_results.csv"
BENCH = BASE / "recursive_functions" / "benchmark"
ABA = "Tempos_Execucao_v4"

# coluna da planilha -> nome base do arquivo (stem) no CSV
COLUNA_PARA_BASE = {
    "B": "fibonacci",
    "C": "factorial",
    "D": "sum",
    "E": "mdc",
    "F": "woodcutter",
    "G": "binary_search",
    "H": "quickselect",
    "I": "identical_strings",
    "J": "linear_search",
    "K": "non_negative_int_contain_digit",
    "L": "Hannoi",
    "M": "MergeSort",
    "N": "NumSearchTree",
    "O": "fast_pow",
    "P": "count_bits",
    "Q": "flatten",
    "R": "first_missing",
    "S": "valid_bst",
    "T": "fib_memo",
    "U": "countdown",
    "V": "safe_parse",
    "W": "sum_nested",
    "X": "remove_keys",
    "Y": "find_first",
    "Z": "count_matches",
}

LINHA_REC = 2
LINHA_TAIL = 3
LINHA_NONREC = 4
LINHA_SOB_TAIL = 5
LINHA_SOB = 6

# Rotulos das linhas de tempo (usados pelas formulas via MATCH na coluna A)
ROTULO_REC = "Tempo_Recursao_Media"
ROTULO_TAIL = "Tempo_Iteracao_Tail_Media"
ROTULO_NONREC = "Tempo_Iteracao_Media"

LABEL_ITER = "Numero_Iteracoes"
LABEL_PARAMS = "Parametros"


def _formula_sobrecarga(rotulo_denominador: str) -> str:
    """Sobrecarga = (linha rec) / (linha denominador), localizadas pelo rotulo.

    INDEX($A:$AA, MATCH(rotulo, $A:$A, 0), COLUMN()) pega o valor na coluna
    atual da linha cujo rotulo (coluna A) casa exatamente. Independe da
    posicao das linhas -> imune a insercao/reordenacao.
    """
    num = f'INDEX($A:$AA,MATCH("{ROTULO_REC}",$A:$A,0),COLUMN())'
    den = f'INDEX($A:$AA,MATCH("{rotulo_denominador}",$A:$A,0),COLUMN())'
    return f'=IFERROR(ROUND({num}/{den},3),"-")'


FORMULA_SOB_TAIL = _formula_sobrecarga(ROTULO_TAIL)
FORMULA_SOB = _formula_sobrecarga(ROTULO_NONREC)


def carregar_csv() -> dict[str, dict]:
    """Mapeia nome_do_arquivo -> linha do CSV."""
    dados = {}
    with open(CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            dados[row["arquivo"]] = row
    return dados


def valor(dados: dict, arquivo: str):
    """Retorna o tempo_ms_por_chamada (float) ou None se nao existir/sem timing."""
    row = dados.get(arquivo)
    if row is None or row["status"] != "ok" or not row["tempo_ms_por_chamada"]:
        return None
    return float(row["tempo_ms_por_chamada"])


def qtd_iter(dados: dict, base: str):
    """qtd_execucoes da funcao (rec, depois tail, depois nonrec). None se ausente."""
    for arq in (f"{base}.py", f"output_{base}.py", f"{base}_nonrec.py"):
        row = dados.get(arq)
        if row and row["status"] == "ok" and row["qtd_execucoes"]:
            return int(row["qtd_execucoes"])
    return None


def achar_ou_inserir_linha(ws, rotulo: str, pos_insercao: int) -> int:
    """Localiza uma linha pelo rotulo na coluna A; insere em pos_insercao se faltar.

    Como as formulas de sobrecarga sao por rotulo (INDEX/MATCH), inserir linhas
    nao as quebra.
    """
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == rotulo:
            return r
    ws.insert_rows(pos_insercao)
    ws.cell(row=pos_insercao, column=1, value=rotulo)
    return pos_insercao


# ----- extracao de parametros dos arquivos de benchmark -----

def _assignments(tree):
    """Mapa nome -> AST do valor das atribuicoes de nivel de modulo."""
    amap = {}
    for node in tree.body:
        if isinstance(node, ast.Assign):
            for t in node.targets:
                if isinstance(t, ast.Name):
                    amap[t.id] = node.value
    return amap


def _substitute(node, amap, depth=0):
    """Substitui Names de modulo pelos seus valores (recursivo, limitado)."""
    if depth > 4:
        return node

    class R(ast.NodeTransformer):
        def visit_Name(self, n):
            if n.id in amap:
                return _substitute(
                    ast.parse(ast.unparse(amap[n.id]), mode="eval").body, amap, depth + 1)
            return n

    return R().visit(node)


def _achar_timeit_lambda(tree):
    """Retorna o ast.Call dentro do lambda passado ao timeit, ou None."""
    for node in ast.walk(tree):
        if isinstance(node, ast.Call):
            f = node.func
            is_timeit = (isinstance(f, ast.Attribute) and f.attr == "timeit") or \
                        (isinstance(f, ast.Name) and f.id == "timeit")
            if not is_timeit:
                continue
            candidatos = list(node.args) + [k.value for k in node.keywords if k.arg == "stmt"]
            for a in candidatos:
                if isinstance(a, ast.Lambda) and isinstance(a.body, ast.Call):
                    return a.body
    return None


def _abreviar_numeros(texto: str) -> str:
    """Encurta literais inteiros longos: 1234...7890(NN dig)."""
    def repl(m):
        s = m.group()
        return f"{s[:6]}...{s[-6:]}({len(s)}dig)"
    return re.sub(r"\d{16,}", repl, texto)


def parametros_de(base: str) -> str:
    """Parametros reais do benchmark recursivo de `base`, formatados 'p1, p2, ...'."""
    arq = BENCH / f"{base}.py"
    if not arq.exists():
        return "-"
    try:
        tree = ast.parse(arq.read_text(encoding="utf-8"))
    except SyntaxError:
        return "-"
    amap = _assignments(tree)
    call = _achar_timeit_lambda(tree)
    if call is None:
        return "-"
    partes = []
    for a in call.args:
        partes.append(_abreviar_numeros(ast.unparse(_substitute(a, amap))))
    for kw in call.keywords:
        partes.append(f"{kw.arg}={_abreviar_numeros(ast.unparse(_substitute(kw.value, amap)))}")
    return ", ".join(partes) if partes else "-"


def main() -> None:
    if not XLSX.exists():
        print(f"Nao encontrado: {XLSX}")
        sys.exit(1)

    dados = carregar_csv()
    wb = openpyxl.load_workbook(XLSX)
    if ABA not in wb.sheetnames:
        print(f"Aba '{ABA}' nao existe. Abas: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[ABA]

    # linhas extra, localizadas/inseridas por rotulo (formulas sao por rotulo)
    linha_iter = achar_ou_inserir_linha(ws, LABEL_ITER, LINHA_SOB + 1)
    linha_params = achar_ou_inserir_linha(ws, LABEL_PARAMS, linha_iter + 1)

    print(f"{'col':>3} {'funcao':<32} {'rec':>10} {'tail':>10} {'nonrec':>10} {'iter':>8}  parametros")
    print("-" * 110)

    for col, base in COLUNA_PARA_BASE.items():
        rec = valor(dados, f"{base}.py")
        nonrec = valor(dados, f"{base}_nonrec.py")
        tail = valor(dados, f"output_{base}.py")
        iteracoes = qtd_iter(dados, base)
        params = parametros_de(base)

        # recursivo (linha 2)
        ws[f"{col}{LINHA_REC}"] = rec if rec is not None else "-"
        # nonrec (linha 4)
        ws[f"{col}{LINHA_NONREC}"] = nonrec if nonrec is not None else "-"
        # tail (linha 3) - so atualiza quem tem output_
        if tail is not None:
            ws[f"{col}{LINHA_TAIL}"] = tail

        # formulas de sobrecarga (por rotulo, robustas a '-')
        ws[f"{col}{LINHA_SOB_TAIL}"] = FORMULA_SOB_TAIL
        ws[f"{col}{LINHA_SOB}"] = FORMULA_SOB

        # numero de iteracoes (qtd_execucoes do benchmark)
        ws[f"{col}{linha_iter}"] = iteracoes if iteracoes is not None else "-"
        # parametros reais do benchmark
        ws[f"{col}{linha_params}"] = params

        srec = f"{rec:.4f}" if rec is not None else "-"
        stail = f"{tail:.4f}" if tail is not None else "(mantido)"
        snon = f"{nonrec:.4f}" if nonrec is not None else "-"
        siter = str(iteracoes) if iteracoes is not None else "-"
        pshow = params if len(params) <= 40 else params[:37] + "..."
        print(f"{col:>3} {base:<32} {srec:>10} {stail:>10} {snon:>10} {siter:>8}  {pshow}")

    wb.save(XLSX)
    print("-" * 110)
    print(f"  iteracoes  -> row {linha_iter}")
    print(f"  parametros -> row {linha_params}")
    print(f"Salvo: {XLSX}")


if __name__ == "__main__":
    main()
