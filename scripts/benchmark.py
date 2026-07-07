"""
Pipeline completo do benchmark, em um unico script e tres fases:

  FASE 1 - VERIFICACAO: para cada funcao, agrupa suas versoes (normal | output_ |
    _nonrec) e confere que usam a MESMA entrada e produzem a MESMA saida. A saida
    e obtida executando a funcao 1x (o timeit e neutralizado -> rapido). Assim
    confirmamos que a comparacao de tempos e justa ANTES de medir.

  FASE 2 - BENCHMARK: roda cada script de verdade (timeit completo), coleta os
    tempos e grava arquivos/csv/benchmark_results.csv.

  FASE 3 - PLANILHA: gera arquivos/xlsx/tempos_execucao.xlsx (aba 'Tempos_Execucao')
    do zero, com toda a estrutura embutida (cabecalhos, complexidade, obs, formulas
    de sobrecarga por rotulo), preenchida com os tempos medidos na fase 2.

Uso:
    python scripts/benchmark.py [diretorio] [--timeout segundos]
    python scripts/benchmark.py --harness <arquivo>   (uso interno da fase 1)

Padrao: recursive_functions/benchmark/
"""

import ast
import contextlib
import csv
import hashlib
import io
import os
import re
import runpy
import subprocess
import sys
import timeit
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).parent.parent
THIS = Path(__file__).resolve()
BENCH = BASE / "recursive_functions" / "benchmark"
CSV_OUT = BASE / "arquivos" / "csv" / "benchmark_results.csv"
XLSX_DIR = BASE / "arquivos" / "xlsx"
XLSX_OUT = XLSX_DIR / "tempos_execucao.xlsx"

SEP = "=" * 70
ABA = "Tempos_Execucao"
ALTURA = 22.5

TIMING_RE = re.compile(
    r"tempo m[eé]dio de (\d+):\s*([\d.]+)s total \| ([\d.]+)ms por chamada"
)

# ordem de exibicao das versoes de cada funcao
ORDEM_TIPO = {"recursivo": 0, "output": 1, "nonrec": 2}


def classify(name: str) -> str:
    if name.startswith("output_"):
        return "output"
    if name.endswith("_nonrec.py"):
        return "nonrec"
    return "recursivo"


def base_de(name: str) -> str:
    """Funcao-base: tira '.py', prefixo 'output_' e sufixo '_nonrec'."""
    stem = name[:-3] if name.endswith(".py") else name
    if stem.startswith("output_"):
        stem = stem[len("output_"):]
    if stem.endswith("_nonrec"):
        stem = stem[:-len("_nonrec")]
    return stem


# ==================== extracao da ENTRADA (AST, estatica) ====================

def _assignments(tree):
    amap = {}
    for node in tree.body:
        if isinstance(node, ast.Assign):
            for t in node.targets:
                if isinstance(t, ast.Name):
                    amap[t.id] = node.value
    return amap


def _substitute(node, amap, depth=0):
    if depth > 4:
        return node

    class R(ast.NodeTransformer):
        def visit_Name(self, n):
            if n.id in amap:
                return _substitute(
                    ast.parse(ast.unparse(amap[n.id]), mode="eval").body, amap, depth + 1)
            return n

    return R().visit(node)


def _achar_timeit_lambda(tree):
    for node in ast.walk(tree):
        if isinstance(node, ast.Call):
            f = node.func
            is_timeit = (isinstance(f, ast.Attribute) and f.attr == "timeit") or \
                        (isinstance(f, ast.Name) and f.id == "timeit")
            if not is_timeit:
                continue
            candidatos = list(node.args) + [k.value for k in node.keywords if k.arg == "stmt"]
            for a in candidatos:
                if isinstance(a, ast.Lambda) and isinstance(a.body, ast.Call):
                    return a.body
    return None


def _abreviar_numeros(texto: str) -> str:
    def repl(m):
        s = m.group()
        return f"{s[:6]}...{s[-6:]}({len(s)}dig)"
    return re.sub(r"\d{16,}", repl, texto)


def entrada_de(path: Path) -> str:
    """Args reais da chamada no timeit, resolvidos e formatados 'p1, p2, ...'."""
    try:
        tree = ast.parse(path.read_text(encoding="utf-8"))
    except (SyntaxError, OSError):
        return "-"
    amap = _assignments(tree)
    call = _achar_timeit_lambda(tree)
    if call is None:
        return "-"
    partes = []
    for a in call.args:
        partes.append(_abreviar_numeros(ast.unparse(_substitute(a, amap))))
    for kw in call.keywords:
        partes.append(f"{kw.arg}={_abreviar_numeros(ast.unparse(_substitute(kw.value, amap)))}")
    return ", ".join(partes) if partes else "-"


# ==================== captura da SAIDA (modo --harness) ====================

def _canonical(v, depth=0) -> str:
    """Representacao canonica SEM endereco de memoria, para comparar saidas."""
    if depth > 8:
        return "..."
    if isinstance(v, ast.AST):
        return ast.dump(v)
    if isinstance(v, (list, tuple)):
        abre, fecha = ("[", "]") if isinstance(v, list) else ("(", ")")
        return abre + ",".join(_canonical(x, depth + 1) for x in v) + fecha
    if isinstance(v, dict):
        return "{" + ",".join(
            f"{_canonical(k, depth + 1)}:{_canonical(val, depth + 1)}" for k, val in v.items()) + "}"
    if isinstance(v, (set, frozenset)):
        return "{" + ",".join(sorted(_canonical(x, depth + 1) for x in v)) + "}"
    r = repr(v)
    if " object at 0x" in r and hasattr(v, "__dict__"):
        return f"{type(v).__name__}({_canonical(vars(v), depth + 1)})"
    return r


def _harness(path_str: str) -> None:
    """Executa `path` com timeit neutralizado (1 chamada) e emite __OUTPUT__."""
    if hasattr(sys, "set_int_max_str_digits"):
        sys.set_int_max_str_digits(0)  # nao limita repr de inteiros gigantes
    path = Path(path_str)
    real = timeit.timeit
    cap = {}

    def patched(stmt="pass", setup="pass", timer=timeit.default_timer,
                number=1000000, globals=None):
        if callable(stmt):
            try:
                cap["out"] = stmt()
            except Exception as e:
                cap["err"] = f"{type(e).__name__}: {e}"
        return 0.0

    timeit.timeit = patched
    buf = io.StringIO()
    try:
        with contextlib.redirect_stdout(buf):
            runpy.run_path(str(path), run_name="__main__")
    except Exception as e:
        print(f"__OUTPUT__\tERRO\t{type(e).__name__}: {e}")
        return
    finally:
        timeit.timeit = real

    if "err" in cap:
        print(f"__OUTPUT__\tERRO\t{cap['err']}")
    elif "out" in cap:
        r = _canonical(cap["out"])
        h = hashlib.md5(r.encode("utf-8", "replace")).hexdigest()[:10]
        preview = r if len(r) <= 70 else r[:67] + "..."
        print(f"__OUTPUT__\t{h}\t{preview.replace(chr(9), ' ').replace(chr(10), ' ')}")
    else:
        print("__OUTPUT__\tn/d\t(sem chamada timeit detectada)")


# ==================== FASE 1: verificacao ====================

def _saida_de(path: Path, timeout: int) -> tuple[str, str]:
    """Roda o harness num subprocesso e devolve (hash, preview) da saida."""
    try:
        proc = subprocess.run(
            [sys.executable, str(THIS), "--harness", str(path)],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
            timeout=timeout,
        )
    except subprocess.TimeoutExpired:
        return "TIMEOUT", f"(>{timeout}s)"
    m = re.search(r"^__OUTPUT__\t(\S+)\t(.*)$", proc.stdout, re.MULTILINE)
    if m:
        return m.group(1), m.group(2)
    ultimo = proc.stderr.strip().splitlines()[-1] if proc.stderr.strip() else "sem saida"
    return "ERRO", ultimo[:70]


def fase_verificacao(files: list[Path], timeout: int) -> int:
    """Confere entrada/saida entre versoes; imprime relatorio. Devolve nº de divergencias."""
    print(f"\n{SEP}\n  FASE 1 - VERIFICACAO (entrada e saida das versoes)\n{SEP}\n")

    resultados = []
    for i, f in enumerate(files, 1):
        print(f"  [{i:>2}/{len(files)}] {f.name}", flush=True)
        h, prev = _saida_de(f, timeout)
        resultados.append({
            "arquivo": f.name, "tipo": classify(f.name), "base": base_de(f.name),
            "entrada": entrada_de(f), "saida_hash": h, "saida_preview": prev,
        })

    por_base = defaultdict(list)
    for r in resultados:
        por_base[r["base"]].append(r)

    print(f"\n{SEP}\n  Comparacao por funcao (normal | output_ | _nonrec)\n{SEP}\n")
    iguais = divergentes = sozinhas = 0
    for base in sorted(por_base):
        versoes = sorted(por_base[base], key=lambda v: ORDEM_TIPO.get(v["tipo"], 9))
        entradas = {v["entrada"] for v in versoes}
        hashes = {v["saida_hash"] for v in versoes}
        bons = all(v["saida_hash"] not in ("ERRO", "TIMEOUT", "n/d") for v in versoes)
        ent_ok = len(entradas) == 1
        sai_ok = bons and len(hashes) == 1

        if len(versoes) < 2:
            sozinhas += 1
            veredito = "(1 versao, nada a comparar)"
        elif ent_ok and sai_ok:
            iguais += 1
            veredito = "OK  (entrada e saida iguais)"
        else:
            divergentes += 1
            sai_txt = "iguais" if sai_ok else ("DIFEREM" if bons else "incompleto")
            veredito = f"DIVERGE  (entrada: {'iguais' if ent_ok else 'DIFEREM'} | saida: {sai_txt})"

        print(f"  {base}  ->  {veredito}")
        if ent_ok:
            print(f"      entrada: {next(iter(entradas))}")
            for v in versoes:
                print(f"        {v['tipo']:<10} {v['arquivo']:<36} saida: {v['saida_hash']} | {v['saida_preview']}")
        else:
            for v in versoes:
                print(f"        {v['tipo']:<10} {v['arquivo']:<36} entrada: {v['entrada']}")
                print(f"        {'':<10} {'':<36} saida  : {v['saida_hash']} | {v['saida_preview']}")
        print()

    print(SEP)
    print(f"  {iguais} OK | {divergentes} divergentes | {sozinhas} sem par")
    print(SEP)
    return divergentes


# ==================== FASE 2: benchmark (timeit real) ====================

def _medir(path: Path, timeout: int) -> dict:
    """Roda o script de verdade e extrai a linha de timing."""
    row = {
        "arquivo": path.name, "tipo": classify(path.name),
        "qtd_execucoes": "", "tempo_total_s": "", "tempo_ms_por_chamada": "", "status": "",
    }
    try:
        # forca o filho a emitir UTF-8 (senao no Windows sai cp1252 e "medio"
        # quebra o regex do timing quando lido como utf-8)
        env = {**os.environ, "PYTHONIOENCODING": "utf-8"}
        proc = subprocess.run(
            [sys.executable, str(path)],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
            timeout=timeout, env=env,
        )
        m = TIMING_RE.search(proc.stdout + proc.stderr)
        if m:
            row.update({
                "qtd_execucoes": int(m.group(1)),
                "tempo_total_s": float(m.group(2)),
                "tempo_ms_por_chamada": float(m.group(3)),
                "status": "ok",
            })
        elif proc.returncode != 0:
            erro = proc.stderr.strip().splitlines()[-1] if proc.stderr.strip() else "erro desconhecido"
            row["status"] = f"erro: {erro[:80]}"
        else:
            row["status"] = "sem_timing"
    except subprocess.TimeoutExpired:
        row["status"] = f"timeout (>{timeout}s)"
    return row


def fase_benchmark(files: list[Path], timeout: int) -> dict[str, dict]:
    """Mede os tempos, grava o CSV e devolve {arquivo: row}."""
    print(f"\n{SEP}\n  FASE 2 - BENCHMARK (timeit completo)  timeout {timeout}s/script\n{SEP}\n")

    dados = {}
    ok = erros = sem_timing = 0
    for i, f in enumerate(files, 1):
        row = _medir(f, timeout)
        dados[f.name] = row
        if row["status"] == "ok":
            print(f"  [{i:>2}/{len(files)}] {f.name:<40} "
                  f"qtd {row['qtd_execucoes']} | {row['tempo_ms_por_chamada']:.4f} ms/chamada")
            ok += 1
        elif row["status"] == "sem_timing":
            print(f"  [{i:>2}/{len(files)}] {f.name:<40} (sem timing)")
            sem_timing += 1
        else:
            print(f"  [{i:>2}/{len(files)}] {f.name:<40} {row['status']}")
            erros += 1

    fieldnames = ["arquivo", "tipo", "qtd_execucoes", "tempo_total_s", "tempo_ms_por_chamada", "status"]
    CSV_OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(CSV_OUT, "w", newline="", encoding="utf-8") as fp:
        w = csv.DictWriter(fp, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(dados.values())

    print(f"\n  {ok} ok | {sem_timing} sem timing | {erros} erros   ->  {CSV_OUT}")
    return dados


# ==================== FASE 3: planilha (xlsx) ====================

COLUNA_PARA_BASE = {
    "B": "fibonacci", "C": "factorial", "D": "sum", "E": "mdc",
    "F": "woodcutter", "G": "binary_search", "H": "quickselect",
    "I": "identical_strings", "J": "linear_search",
    "K": "non_negative_int_contain_digit", "L": "Hannoi", "M": "MergeSort",
    "N": "NumSearchTree", "O": "fast_pow", "P": "count_bits", "Q": "flatten",
    "R": "first_missing", "S": "valid_bst", "T": "fib_memo", "U": "countdown",
    "V": "safe_parse", "W": "sum_nested", "X": "remove_keys", "Y": "find_first",
    "Z": "count_matches",
}

# Cabecalho por funcao (linha 1): apenas o NOME, sem parametros -- os
# parametros exatos ficam na linha 'Parametros'.
HEADERS = {
    "B": "Fibonacci", "C": "Factorial", "D": "Sum", "E": "mdc",
    "F": "woodcutter", "G": "Binary_Search_Tree", "H": "quick_select",
    "I": "identical_strings", "J": "linear_search",
    "K": "non_negative_int_contain_digit", "L": "Hannoi", "M": "MergeSort",
    "N": "NumSearchTree", "O": "fast_pow", "P": "count_bits", "Q": "flatten",
    "R": "first_missing", "S": "is_valid_bst", "T": "fib_memo", "U": "countdown",
    "V": "safe_parse", "W": "sum_nested", "X": "remove_keys", "Y": "find_first",
    "Z": "count_matches",
}

COMPLEXIDADE = {
    "B": "O(n)", "C": "O(n)", "D": "O(n)", "E": "O(log min(a,b))",
    "F": "O(n log m)", "G": "O(log n)", "H": "O(n) médio / O(n²) pior caso",
    "I": "O(n)", "J": "O(n)", "K": "O(log n)", "L": "O(2ⁿ)", "M": "O(n log n)",
    "N": "O(4ⁿ / n^1.5)", "O": "O(log exp)", "P": "O(log n)", "Q": "O(n²)",
    "R": "O(n²)", "S": "O(n)", "T": "O(n)", "U": "O(n)", "V": "O(n²)",
    "W": "O(n)", "X": "O(k²)", "Y": "O(n)", "Z": "O(n²)",
}

OBS = {
    "F": "n = nº de árvores, m = intervalo de altura",
    "H": "Depende da escolha do pivô",
    "I": "n = tamanho da string",
    "L": "Gera 2ⁿ − 1 movimentos obrigatoriamente",
    "O": "exp reduz pela metade a cada dois passos",
}

LINHA_REC, LINHA_TAIL, LINHA_NONREC = 2, 3, 4
LINHA_SOB_TAIL, LINHA_SOB = 5, 6
LINHA_ITER, LINHA_PARAMS, LINHA_CPX, LINHA_OBS = 7, 8, 9, 10
ULTIMA_LINHA = LINHA_OBS

ROTULO_REC = "Tempo_Recursao_Media"
ROTULO_TAIL = "Tempo_Iteracao_Tail_Media"
ROTULO_NONREC = "Tempo_Iteracao_Media"

ROTULOS = {
    LINHA_REC: ROTULO_REC, LINHA_TAIL: ROTULO_TAIL, LINHA_NONREC: ROTULO_NONREC,
    LINHA_SOB_TAIL: "Sobrecarga_Tail (rec/it_tail)",
    LINHA_SOB: "Sobrecarga (rec/it_normal)",
    LINHA_ITER: "Numero_Iteracoes", LINHA_PARAMS: "Parametros",
    LINHA_CPX: "Complexidade de resolução", LINHA_OBS: "Obs",
}

HDR_FILL = PatternFill(start_color="FF217346", end_color="FF217346", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFFFF")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BAND_FILL = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
ROT_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
ALINHA_LINHA = {LINHA_PARAMS: LEFT, LINHA_OBS: LEFT}


def _formula_sobrecarga(rotulo_denominador: str) -> str:
    num = f'INDEX($A:$AA,MATCH("{ROTULO_REC}",$A:$A,0),COLUMN())'
    den = f'INDEX($A:$AA,MATCH("{rotulo_denominador}",$A:$A,0),COLUMN())'
    return f'=IFERROR(ROUND({num}/{den},3),"-")'


FORMULA_SOB_TAIL = _formula_sobrecarga(ROTULO_TAIL)
FORMULA_SOB = _formula_sobrecarga(ROTULO_NONREC)


def _tempo(dados: dict, arquivo: str):
    """tempo_ms_por_chamada (float) ou None se ausente/sem timing."""
    row = dados.get(arquivo)
    if row is None or row["status"] != "ok" or not row["tempo_ms_por_chamada"]:
        return None
    return float(row["tempo_ms_por_chamada"])


def _qtd_iter(dados: dict, base: str):
    for arq in (f"{base}.py", f"output_{base}.py", f"{base}_nonrec.py"):
        row = dados.get(arq)
        if row and row["status"] == "ok" and row["qtd_execucoes"]:
            return int(row["qtd_execucoes"])
    return None


def fase_planilha(dados: dict[str, dict]) -> None:
    """Gera o xlsx do zero a partir dos tempos medidos (dict {arquivo: row})."""
    print(f"\n{SEP}\n  FASE 3 - PLANILHA (tempos_execucao.xlsx)\n{SEP}\n")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ABA

    ws.cell(1, 1, "Função")
    for linha, rotulo in ROTULOS.items():
        c = ws.cell(linha, 1, rotulo)
        c.font = ROT_FONT
        c.alignment = LEFT

    for col, base in COLUNA_PARA_BASE.items():
        rec = _tempo(dados, f"{base}.py")
        nonrec = _tempo(dados, f"{base}_nonrec.py")
        tail = _tempo(dados, f"output_{base}.py")
        iteracoes = _qtd_iter(dados, base)
        params = entrada_de(BENCH / f"{base}.py")

        ws[f"{col}1"] = HEADERS.get(col, base)
        ws[f"{col}{LINHA_REC}"] = rec if rec is not None else "-"
        ws[f"{col}{LINHA_NONREC}"] = nonrec if nonrec is not None else "-"
        ws[f"{col}{LINHA_TAIL}"] = tail if tail is not None else "-"
        ws[f"{col}{LINHA_SOB_TAIL}"] = FORMULA_SOB_TAIL
        ws[f"{col}{LINHA_SOB}"] = FORMULA_SOB
        ws[f"{col}{LINHA_ITER}"] = iteracoes if iteracoes is not None else "-"
        ws[f"{col}{LINHA_PARAMS}"] = params
        ws[f"{col}{LINHA_CPX}"] = COMPLEXIDADE.get(col, "-")
        ws[f"{col}{LINHA_OBS}"] = OBS.get(col, "")
        for linha in range(LINHA_REC, ULTIMA_LINHA + 1):
            ws[f"{col}{linha}"].alignment = ALINHA_LINHA.get(linha, CENTER)

    ncols = 1 + len(COLUNA_PARA_BASE)
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = HDR_ALIGN
    for r in range(LINHA_REC, ULTIMA_LINHA + 1):
        if r % 2 == 0:
            for c in range(1, ncols + 1):
                ws.cell(r, c).fill = BAND_FILL
    for r in range(1, ULTIMA_LINHA + 1):
        ws.row_dimensions[r].height = ALTURA
    ws.column_dimensions["A"].width = 30
    for col in COLUNA_PARA_BASE:
        ws.column_dimensions[col].width = 16
    ws.freeze_panes = "B2"

    XLSX_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_OUT)
    print(f"  Aba '{ABA}' com {len(COLUNA_PARA_BASE)} funcoes  ->  {XLSX_OUT}")


# ==================== orquestracao ====================

def main() -> None:
    args = sys.argv[1:]
    directory = BENCH
    timeout = 120
    i = 0
    while i < len(args):
        if args[i] == "--timeout" and i + 1 < len(args):
            timeout = int(args[i + 1])
            i += 2
        else:
            directory = Path(args[i])
            if not directory.is_absolute():
                directory = BASE / directory
            i += 1

    if not directory.is_dir():
        print(f"Diretorio nao encontrado: {directory}")
        sys.exit(1)

    files = sorted(f for f in directory.iterdir() if f.is_file() and f.suffix == ".py")
    if not files:
        print(f"Nenhum .py em: {directory}")
        sys.exit(1)

    print(f"\n{SEP}\n  BENCHMARK - {len(files)} arquivos em {directory}\n{SEP}")

    divergentes = fase_verificacao(files, timeout)
    if divergentes:
        print(f"\n  [AVISO] {divergentes} funcao(oes) com entrada/saida divergente entre "
              f"versoes: a comparacao de tempos delas pode nao ser justa.\n")

    dados = fase_benchmark(files, timeout)
    fase_planilha(dados)

    print(f"\n{SEP}\n  Concluido: verificacao + benchmark + planilha\n{SEP}")


if __name__ == "__main__":
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    if len(sys.argv) >= 3 and sys.argv[1] == "--harness":
        _harness(sys.argv[2])
        sys.exit(0)

    main()
