"""
Gera arquivos/xlsx/tempos_execucao.xlsx (aba 'Tempos_Execucao') a partir de
arquivos/csv/benchmark_results.csv, produzido pelo benchmark.py no modo
classico. Nada e remedido: toda a estrutura da planilha (cabecalhos,
complexidade, obs, formulas de sobrecarga por rotulo) e embutida aqui e
preenchida com os tempos ja gravados no CSV.

Uso:
    python scripts/planilha_tempos.py
"""

import csv
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from benchmark import BENCH, CSV_OUT, entrada_de

XLSX_DIR = CSV_OUT.parent.parent / "xlsx"
XLSX_OUT = XLSX_DIR / "tempos_execucao.xlsx"

ABA = "Tempos_Execucao"
ALTURA = 22.5

COLUNA_PARA_BASE = {
    "B": "fibonacci", "C": "factorial", "D": "sum", "E": "mdc",
    "F": "woodcutter", "G": "binary_search", "H": "quickselect",
    "I": "identical_strings", "J": "linear_search",
    "K": "non_negative_int_contain_digit", "L": "Hannoi", "M": "MergeSort",
    "N": "NumSearchTree", "O": "fast_pow", "P": "count_bits", "Q": "flatten",
    "R": "first_missing", "S": "valid_bst", "T": "fib_memo", "U": "countdown",
    "V": "safe_parse", "W": "sum_nested", "X": "remove_keys", "Y": "find_first",
    "Z": "count_matches",
}

# Cabecalho por funcao (linha 1): apenas o NOME, sem parametros -- os
# parametros exatos ficam na linha 'Parametros'.
HEADERS = {
    "B": "Fibonacci", "C": "Factorial", "D": "Sum", "E": "mdc",
    "F": "woodcutter", "G": "Binary_Search_Tree", "H": "quick_select",
    "I": "identical_strings", "J": "linear_search",
    "K": "non_negative_int_contain_digit", "L": "Hannoi", "M": "MergeSort",
    "N": "NumSearchTree", "O": "fast_pow", "P": "count_bits", "Q": "flatten",
    "R": "first_missing", "S": "is_valid_bst", "T": "fib_memo", "U": "countdown",
    "V": "safe_parse", "W": "sum_nested", "X": "remove_keys", "Y": "find_first",
    "Z": "count_matches",
}

COMPLEXIDADE = {
    "B": "O(n)", "C": "O(n)", "D": "O(n)", "E": "O(log min(a,b))",
    "F": "O(n log m)", "G": "O(log n)", "H": "O(n) médio / O(n²) pior caso",
    "I": "O(n)", "J": "O(n)", "K": "O(log n)", "L": "O(2ⁿ)", "M": "O(n log n)",
    "N": "O(4ⁿ / n^1.5)", "O": "O(log exp)", "P": "O(log n)", "Q": "O(n²)",
    "R": "O(n²)", "S": "O(n)", "T": "O(n)", "U": "O(n)", "V": "O(n²)",
    "W": "O(n)", "X": "O(k²)", "Y": "O(n)", "Z": "O(n²)",
}

OBS = {
    "F": "n = nº de árvores, m = intervalo de altura",
    "H": "Depende da escolha do pivô",
    "I": "n = tamanho da string",
    "L": "Gera 2ⁿ − 1 movimentos obrigatoriamente",
    "O": "exp reduz pela metade a cada dois passos",
}

LINHA_REC, LINHA_TAIL, LINHA_NONREC = 2, 3, 4
LINHA_SOB_TAIL, LINHA_SOB = 5, 6
LINHA_ITER, LINHA_PARAMS, LINHA_CPX, LINHA_OBS = 7, 8, 9, 10
ULTIMA_LINHA = LINHA_OBS

ROTULO_REC = "Tempo_Recursao_Media"
ROTULO_TAIL = "Tempo_Iteracao_Tail_Media"
ROTULO_NONREC = "Tempo_Iteracao_Media"

ROTULOS = {
    LINHA_REC: ROTULO_REC, LINHA_TAIL: ROTULO_TAIL, LINHA_NONREC: ROTULO_NONREC,
    LINHA_SOB_TAIL: "Sobrecarga_Tail (rec/it_tail)",
    LINHA_SOB: "Sobrecarga (rec/it_normal)",
    LINHA_ITER: "Numero_Iteracoes", LINHA_PARAMS: "Parametros",
    LINHA_CPX: "Complexidade de resolução", LINHA_OBS: "Obs",
}

HDR_FILL = PatternFill(start_color="FF217346", end_color="FF217346", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFFFF")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BAND_FILL = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
ROT_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
ALINHA_LINHA = {LINHA_PARAMS: LEFT, LINHA_OBS: LEFT}


def _formula_sobrecarga(rotulo_denominador: str) -> str:
    num = f'INDEX($A:$AA,MATCH("{ROTULO_REC}",$A:$A,0),COLUMN())'
    den = f'INDEX($A:$AA,MATCH("{rotulo_denominador}",$A:$A,0),COLUMN())'
    return f'=IFERROR(ROUND({num}/{den},3),"-")'


FORMULA_SOB_TAIL = _formula_sobrecarga(ROTULO_TAIL)
FORMULA_SOB = _formula_sobrecarga(ROTULO_NONREC)


def _tempo(dados: dict, arquivo: str):
    """tempo_ms_por_chamada (float) ou None se ausente/sem timing."""
    row = dados.get(arquivo)
    if row is None or row["status"] != "ok" or not row["tempo_ms_por_chamada"]:
        return None
    return float(row["tempo_ms_por_chamada"])


def _qtd_iter(dados: dict, base: str):
    for arq in (f"{base}.py", f"output_{base}.py", f"{base}_nonrec.py"):
        row = dados.get(arq)
        if row and row["status"] == "ok" and row["qtd_execucoes"]:
            return int(row["qtd_execucoes"])
    return None


def gerar_planilha(dados: dict[str, dict]) -> None:
    """Gera o xlsx do zero a partir dos tempos medidos (dict {arquivo: row})."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ABA

    ws.cell(1, 1, "Função")
    for linha, rotulo in ROTULOS.items():
        c = ws.cell(linha, 1, rotulo)
        c.font = ROT_FONT
        c.alignment = LEFT

    for col, base in COLUNA_PARA_BASE.items():
        rec = _tempo(dados, f"{base}.py")
        nonrec = _tempo(dados, f"{base}_nonrec.py")
        tail = _tempo(dados, f"output_{base}.py")
        iteracoes = _qtd_iter(dados, base)
        params = entrada_de(BENCH / f"{base}.py")

        ws[f"{col}1"] = HEADERS.get(col, base)
        ws[f"{col}{LINHA_REC}"] = rec if rec is not None else "-"
        ws[f"{col}{LINHA_NONREC}"] = nonrec if nonrec is not None else "-"
        ws[f"{col}{LINHA_TAIL}"] = tail if tail is not None else "-"
        ws[f"{col}{LINHA_SOB_TAIL}"] = FORMULA_SOB_TAIL
        ws[f"{col}{LINHA_SOB}"] = FORMULA_SOB
        ws[f"{col}{LINHA_ITER}"] = iteracoes if iteracoes is not None else "-"
        ws[f"{col}{LINHA_PARAMS}"] = params
        ws[f"{col}{LINHA_CPX}"] = COMPLEXIDADE.get(col, "-")
        ws[f"{col}{LINHA_OBS}"] = OBS.get(col, "")
        for linha in range(LINHA_REC, ULTIMA_LINHA + 1):
            ws[f"{col}{linha}"].alignment = ALINHA_LINHA.get(linha, CENTER)

    ncols = 1 + len(COLUNA_PARA_BASE)
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = HDR_ALIGN
    for r in range(LINHA_REC, ULTIMA_LINHA + 1):
        if r % 2 == 0:
            for c in range(1, ncols + 1):
                ws.cell(r, c).fill = BAND_FILL
    for r in range(1, ULTIMA_LINHA + 1):
        ws.row_dimensions[r].height = ALTURA
    ws.column_dimensions["A"].width = 30
    for col in COLUNA_PARA_BASE:
        ws.column_dimensions[col].width = 16
    ws.freeze_panes = "B2"

    XLSX_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_OUT)
    print(f"  Aba '{ABA}' com {len(COLUNA_PARA_BASE)} funcoes  ->  {XLSX_OUT}")


def main() -> None:
    if not CSV_OUT.exists():
        print(f"CSV nao encontrado: {CSV_OUT} (rode scripts/benchmark.py antes)")
        sys.exit(1)
    with open(CSV_OUT, newline="", encoding="utf-8") as fp:
        dados = {row["arquivo"]: row for row in csv.DictReader(fp)}
    gerar_planilha(dados)


if __name__ == "__main__":
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    main()
