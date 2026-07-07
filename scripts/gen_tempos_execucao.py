"""
Cria do zero o arquivos/xlsx/tempos_execucao.xlsx com a pagina 'Tempos_Execucao'.

NAO depende de nenhum xlsx template (ling_rec.xlsx / backup). Toda a estrutura
antes herdada da aba 'Tempos_Execucao_v4' agora esta embutida como constantes
editaveis: os cabecalhos descritivos (HEADERS), a complexidade (COMPLEXIDADE) e
as observacoes (OBS). O visual (cabecalho verde + zebra) e aplicado manualmente,
sem Tabela/ListObject.

Linhas (rotulo na coluna A):
    1  Funcao                     <- cabecalho descritivo por funcao (HEADERS)
    2  Tempo_Recursao_Media       <- <nome>.py           (recursivo)
    3  Tempo_Iteracao_Tail_Media  <- output_<nome>.py; '-' se nao ha versao tail
    4  Tempo_Iteracao_Media       <- <nome>_nonrec.py    (nonrec)
    5  Sobrecarga_Tail            = formula rec/tail   (por rotulo)
    6  Sobrecarga                 = formula rec/nonrec (por rotulo)
    7  Numero_Iteracoes           <- qtd_execucoes do benchmark
    8  Parametros                 <- args reais no timeit(lambda: f(...))
    9  Complexidade de resolucao  <- COMPLEXIDADE (constante)
    10 Obs                        <- OBS (constante)

Fontes de dados: arquivos/csv/benchmark_results.csv (tempos) e os .py do
benchmark (parametros). Metrica: tempo_ms_por_chamada. Ausentes viram '-'.

Uso:
    python scripts/gen_tempos_execucao.py
"""

import ast
import csv
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).parent.parent
SAIDA_DIR = BASE / "arquivos" / "xlsx"  # diretorio padrao de saida dos xlsx gerados
SAIDA = SAIDA_DIR / "tempos_execucao.xlsx"
CSV = BASE / "arquivos" / "csv" / "benchmark_results.csv"
BENCH = BASE / "recursive_functions" / "benchmark"
ABA = "Tempos_Execucao"
ALTURA = 22.5

# ---- estilo ----
HDR_FILL = PatternFill(start_color="FF217346", end_color="FF217346", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFFFF")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BAND_FILL = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
ROT_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

# coluna da planilha -> nome base do arquivo (stem) no CSV
COLUNA_PARA_BASE = {
    "B": "fibonacci", "C": "factorial", "D": "sum", "E": "mdc",
    "F": "woodcutter", "G": "binary_search", "H": "quickselect",
    "I": "identical_strings", "J": "linear_search",
    "K": "non_negative_int_contain_digit", "L": "Hannoi", "M": "MergeSort",
    "N": "NumSearchTree", "O": "fast_pow", "P": "count_bits", "Q": "flatten",
    "R": "first_missing", "S": "valid_bst", "T": "fib_memo", "U": "countdown",
    "V": "safe_parse", "W": "sum_nested", "X": "remove_keys", "Y": "find_first",
    "Z": "count_matches",
}

# ---- conteudo manual embutido (antes vinha do template) ----
# Cabecalho descritivo por funcao (linha 1). Sao rotulos humanos; os parametros
# exatos ficam na linha 'Parametros'. Edite aqui se quiser mudar o rotulo.
HEADERS = {
    "B": "Fibonaccil(500.000)",
    "C": "Factorial(50.000, 1)",
    "D": "Sum (5.000.000)",
    "E": "mdc(fib(500), fib(501))",
    "F": "woodcutter([1, ..., 100001), 1250025000, 0, max(arvores))",
    "G": "Binary_Search_Tree()",
    "H": "quick_select()",
    "I": "identical_strings()",
    "J": "linear_search()",
    "K": "non_negative_int_contain_digit()",
    "L": "Hannoi (20)",
    "M": "MergeSort(1.000.000)",
    "N": "NumSearchTree (18)",
    "O": "fast_pow(2, 10.000)",
    "P": "count_bits((1 << 20) - 1)",
    "Q": "flatten()",
    "R": "first_missing(range(500))",
    "S": "is_valid_bst(1.000)",
    "T": "fib_memo(35)",
    "U": "countdown(500)",
    "V": "safe_parse(['1','abc','3',None,'5'] * 200)",
    "W": "sum_nested()",
    "X": "remove_keys(dict)",
    "Y": "find_first(range(1.000), x > 5000)",
    "Z": "count_matches([1,2,3] * 333, 2)",
}

COMPLEXIDADE = {
    "B": "O(n)", "C": "O(n)", "D": "O(n)", "E": "O(log min(a,b))",
    "F": "O(n log m)", "G": "O(log n)", "H": "O(n) médio / O(n²) pior caso",
    "I": "O(n)", "J": "O(n)", "K": "O(log n)", "L": "O(2ⁿ)", "M": "O(n log n)",
    "N": "O(4ⁿ / n^1.5)", "O": "O(log exp)", "P": "O(log n)", "Q": "O(n²)",
    "R": "O(n²)", "S": "O(n)", "T": "O(n)", "U": "O(n)", "V": "O(n²)",
    "W": "O(n)", "X": "O(k²)", "Y": "O(n)", "Z": "O(n²)",
}

OBS = {
    "F": "n = nº de árvores, m = intervalo de altura",
    "H": "Depende da escolha do pivô",
    "I": "n = tamanho da string",
    "L": "Gera 2ⁿ − 1 movimentos obrigatoriamente",
    "O": "exp reduz pela metade a cada dois passos",
}

# ---- layout das linhas ----
LINHA_REC = 2
LINHA_TAIL = 3
LINHA_NONREC = 4
LINHA_SOB_TAIL = 5
LINHA_SOB = 6
LINHA_ITER = 7
LINHA_PARAMS = 8
LINHA_CPX = 9
LINHA_OBS = 10
ULTIMA_LINHA = LINHA_OBS

# Rotulos das linhas de tempo (usados pelas formulas via MATCH na coluna A)
ROTULO_REC = "Tempo_Recursao_Media"
ROTULO_TAIL = "Tempo_Iteracao_Tail_Media"
ROTULO_NONREC = "Tempo_Iteracao_Media"

ROTULOS = {
    LINHA_REC: ROTULO_REC,
    LINHA_TAIL: ROTULO_TAIL,
    LINHA_NONREC: ROTULO_NONREC,
    LINHA_SOB_TAIL: "Sobrecarga_Tail (rec/it_tail)",
    LINHA_SOB: "Sobrecarga (rec/it_normal)",
    LINHA_ITER: "Numero_Iteracoes",
    LINHA_PARAMS: "Parametros",
    LINHA_CPX: "Complexidade de resolução",
    LINHA_OBS: "Obs",
}

# alinhamento das celulas de valor por linha (default: centro)
ALINHA_LINHA = {LINHA_PARAMS: LEFT, LINHA_OBS: LEFT}


def _formula_sobrecarga(rotulo_denominador: str) -> str:
    """Sobrecarga = (linha rec) / (linha denominador), localizadas pelo rotulo.

    INDEX($A:$AA, MATCH(rotulo, $A:$A, 0), COLUMN()) pega o valor na coluna
    atual da linha cujo rotulo (coluna A) casa exatamente -> imune a
    insercao/reordenacao de linhas.
    """
    num = f'INDEX($A:$AA,MATCH("{ROTULO_REC}",$A:$A,0),COLUMN())'
    den = f'INDEX($A:$AA,MATCH("{rotulo_denominador}",$A:$A,0),COLUMN())'
    return f'=IFERROR(ROUND({num}/{den},3),"-")'


FORMULA_SOB_TAIL = _formula_sobrecarga(ROTULO_TAIL)
FORMULA_SOB = _formula_sobrecarga(ROTULO_NONREC)


def carregar_csv() -> dict[str, dict]:
    """Mapeia nome_do_arquivo -> linha do CSV."""
    dados = {}
    with open(CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            dados[row["arquivo"]] = row
    return dados


def valor(dados: dict, arquivo: str):
    """Retorna o tempo_ms_por_chamada (float) ou None se nao existir/sem timing."""
    row = dados.get(arquivo)
    if row is None or row["status"] != "ok" or not row["tempo_ms_por_chamada"]:
        return None
    return float(row["tempo_ms_por_chamada"])


def qtd_iter(dados: dict, base: str):
    """qtd_execucoes da funcao (rec, depois tail, depois nonrec). None se ausente."""
    for arq in (f"{base}.py", f"output_{base}.py", f"{base}_nonrec.py"):
        row = dados.get(arq)
        if row and row["status"] == "ok" and row["qtd_execucoes"]:
            return int(row["qtd_execucoes"])
    return None


# ----- extracao de parametros dos arquivos de benchmark -----

def _assignments(tree):
    """Mapa nome -> AST do valor das atribuicoes de nivel de modulo."""
    amap = {}
    for node in tree.body:
        if isinstance(node, ast.Assign):
            for t in node.targets:
                if isinstance(t, ast.Name):
                    amap[t.id] = node.value
    return amap


def _substitute(node, amap, depth=0):
    """Substitui Names de modulo pelos seus valores (recursivo, limitado)."""
    if depth > 4:
        return node

    class R(ast.NodeTransformer):
        def visit_Name(self, n):
            if n.id in amap:
                return _substitute(
                    ast.parse(ast.unparse(amap[n.id]), mode="eval").body, amap, depth + 1)
            return n

    return R().visit(node)


def _achar_timeit_lambda(tree):
    """Retorna o ast.Call dentro do lambda passado ao timeit, ou None."""
    for node in ast.walk(tree):
        if isinstance(node, ast.Call):
            f = node.func
            is_timeit = (isinstance(f, ast.Attribute) and f.attr == "timeit") or \
                        (isinstance(f, ast.Name) and f.id == "timeit")
            if not is_timeit:
                continue
            candidatos = list(node.args) + [k.value for k in node.keywords if k.arg == "stmt"]
            for a in candidatos:
                if isinstance(a, ast.Lambda) and isinstance(a.body, ast.Call):
                    return a.body
    return None


def _abreviar_numeros(texto: str) -> str:
    """Encurta literais inteiros longos: 1234...7890(NN dig)."""
    def repl(m):
        s = m.group()
        return f"{s[:6]}...{s[-6:]}({len(s)}dig)"
    return re.sub(r"\d{16,}", repl, texto)


def parametros_de(base: str) -> str:
    """Parametros reais do benchmark recursivo de `base`, formatados 'p1, p2, ...'."""
    arq = BENCH / f"{base}.py"
    if not arq.exists():
        return "-"
    try:
        tree = ast.parse(arq.read_text(encoding="utf-8"))
    except SyntaxError:
        return "-"
    amap = _assignments(tree)
    call = _achar_timeit_lambda(tree)
    if call is None:
        return "-"
    partes = []
    for a in call.args:
        partes.append(_abreviar_numeros(ast.unparse(_substitute(a, amap))))
    for kw in call.keywords:
        partes.append(f"{kw.arg}={_abreviar_numeros(ast.unparse(_substitute(kw.value, amap)))}")
    return ", ".join(partes) if partes else "-"


def main() -> None:
    dados = carregar_csv()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ABA

    # coluna A: canto + rotulos das linhas
    ws.cell(1, 1, "Função")
    for linha, rotulo in ROTULOS.items():
        c = ws.cell(linha, 1, rotulo)
        c.font = ROT_FONT
        c.alignment = LEFT

    print(f"{'col':>3} {'funcao':<32} {'rec':>10} {'tail':>10} {'nonrec':>10} {'iter':>8}  parametros")
    print("-" * 110)

    for col, base in COLUNA_PARA_BASE.items():
        rec = valor(dados, f"{base}.py")
        nonrec = valor(dados, f"{base}_nonrec.py")
        tail = valor(dados, f"output_{base}.py")
        iteracoes = qtd_iter(dados, base)
        params = parametros_de(base)

        ws[f"{col}1"] = HEADERS.get(col, base)
        ws[f"{col}{LINHA_REC}"] = rec if rec is not None else "-"
        ws[f"{col}{LINHA_NONREC}"] = nonrec if nonrec is not None else "-"
        # tail: valor medido, ou '-' quando nao ha output_ (sem versao tail-recursiva)
        ws[f"{col}{LINHA_TAIL}"] = tail if tail is not None else "-"
        ws[f"{col}{LINHA_SOB_TAIL}"] = FORMULA_SOB_TAIL
        ws[f"{col}{LINHA_SOB}"] = FORMULA_SOB
        ws[f"{col}{LINHA_ITER}"] = iteracoes if iteracoes is not None else "-"
        ws[f"{col}{LINHA_PARAMS}"] = params
        ws[f"{col}{LINHA_CPX}"] = COMPLEXIDADE.get(col, "-")
        ws[f"{col}{LINHA_OBS}"] = OBS.get(col, "")

        # alinhamento das celulas de valor
        for linha in range(LINHA_REC, ULTIMA_LINHA + 1):
            ws[f"{col}{linha}"].alignment = ALINHA_LINHA.get(linha, CENTER)

        srec = f"{rec:.4f}" if rec is not None else "-"
        stail = f"{tail:.4f}" if tail is not None else "-"
        snon = f"{nonrec:.4f}" if nonrec is not None else "-"
        siter = str(iteracoes) if iteracoes is not None else "-"
        pshow = params if len(params) <= 40 else params[:37] + "..."
        print(f"{col:>3} {base:<32} {srec:>10} {stail:>10} {snon:>10} {siter:>8}  {pshow}")

    # estilo do cabecalho (linha 1)
    ncols = 1 + len(COLUNA_PARA_BASE)
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = HDR_ALIGN

    # zebra nas linhas de dados (banding manual, no lugar do estilo de tabela)
    for r in range(LINHA_REC, ULTIMA_LINHA + 1):
        if r % 2 == 0:
            for c in range(1, ncols + 1):
                ws.cell(r, c).fill = BAND_FILL

    # altura uniforme, larguras e paineis congelados
    for r in range(1, ULTIMA_LINHA + 1):
        ws.row_dimensions[r].height = ALTURA
    ws.column_dimensions["A"].width = 30
    for col in COLUNA_PARA_BASE:
        ws.column_dimensions[col].width = 16
    ws.freeze_panes = "B2"

    SAIDA_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(SAIDA)

    print("-" * 110)
    print(f"Arquivo criado : {SAIDA}")
    print(f"Aba            : {ABA} (gerada do zero, sem template)")
    print(f"Funcoes        : {len(COLUNA_PARA_BASE)} (colunas B..{list(COLUNA_PARA_BASE)[-1]})")
    print(f"Linhas         : 1..{ULTIMA_LINHA} (sem 'notacao')")


if __name__ == "__main__":
    main()
